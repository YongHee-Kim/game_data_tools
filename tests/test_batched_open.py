"""Verify that multi-sheet export/import open and save the xlsx exactly once."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl
from openpyxl import Workbook

from game_data_tools import Project
from game_data_tools.workbook import JSONWorkbook


def _build_multisheet_project(root: Path) -> None:
    xlsx_dir = root / "xlsx"
    out_dir = root / "json"
    xlsx_dir.mkdir()
    out_dir.mkdir()

    wb = Workbook()
    wb.remove(wb.active)
    for name in ("Weapons", "Armor", "Consumables"):
        ws = wb.create_sheet(name)
        ws.append(["Key", "Name", "/stats/value"])
        ws.append([f"{name}001", f"{name} A", 10])
        ws.append([f"{name}002", f"{name} B", 20])
    wb.save(xlsx_dir / "Items.xlsx")

    (root / "config.json").write_text(
        json.dumps(
            {
                "name": "Batch",
                "environment": {"xlsx": "./xlsx", "out": "./json"},
                "xlsxtables": {
                    "Items.xlsx": {
                        "workSheets": [
                            {"name": "Weapons", "out": "Items_Weapons.json"},
                            {"name": "Armor", "out": "Items_Armor.json"},
                            {"name": "Consumables", "out": "Items_Consumables.json"},
                        ]
                    }
                },
            }
        ),
        encoding="utf-8",
    )


class BatchedOpenTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        _build_multisheet_project(self.root)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_export_opens_workbook_once(self) -> None:
        project = Project(self.root)
        real_load = openpyxl.load_workbook
        calls: list[Path] = []

        def counting_loader(path, *args, **kwargs):
            calls.append(Path(path))
            return real_load(path, *args, **kwargs)

        with mock.patch("game_data_tools.workbook.load_workbook", side_effect=counting_loader):
            results = project.export()

        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].errors, [])
        self.assertEqual(len(results[0].written), 3)
        self.assertEqual(len(calls), 1, f"expected 1 open, got {len(calls)}: {calls}")

    def test_import_opens_workbook_once(self) -> None:
        project = Project(self.root)
        project.export()  # produce the json files

        real_load = openpyxl.load_workbook
        load_calls: list[Path] = []
        save_calls: list[Path] = []

        def counting_loader(path, *args, **kwargs):
            load_calls.append(Path(path))
            return real_load(path, *args, **kwargs)

        real_save = Workbook.save

        def counting_save(self, path):
            save_calls.append(Path(path))
            return real_save(self, path)

        with mock.patch("game_data_tools.workbook.load_workbook", side_effect=counting_loader), \
             mock.patch.object(Workbook, "save", counting_save):
            result = project.import_json("items")

        self.assertEqual(result.errors, [])
        self.assertEqual(result.sheets, ["Weapons", "Armor", "Consumables"])
        self.assertEqual(len(load_calls), 1, f"expected 1 open, got {len(load_calls)}")
        self.assertEqual(len(save_calls), 1, f"expected 1 save, got {len(save_calls)}")

    def test_import_json_skips_save_when_all_sheets_fail(self) -> None:
        project = Project(self.root)
        # No JSON files exist on disk yet — every sheet's from_json call will fail.
        xlsx_path = self.root / "xlsx" / "Items.xlsx"
        original_bytes = xlsx_path.read_bytes()

        save_calls: list[Path] = []
        real_save = Workbook.save

        def counting_save(self, path):
            save_calls.append(Path(path))
            return real_save(self, path)

        with mock.patch.object(Workbook, "save", counting_save):
            result = project.import_json("items")

        self.assertEqual(result.sheets, [])
        self.assertEqual(len(result.errors), 3)
        self.assertEqual(save_calls, [], "import_json must not save when no sheets succeeded")
        self.assertEqual(xlsx_path.read_bytes(), original_bytes)

    def test_jsonworkbook_context_manager_closes_on_exception(self) -> None:
        fake_wb = mock.MagicMock()
        jwb = JSONWorkbook(fake_wb, Path("dummy.xlsx"))

        with self.assertRaises(RuntimeError):
            with jwb:
                raise RuntimeError("boom")

        fake_wb.close.assert_called_once()


if __name__ == "__main__":
    unittest.main()
