"""xlsx -> json -> xlsx -> json round trip: the second json must equal the first."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from game_data_tools import JSONWorksheet, Project


def _build_project(root: Path) -> None:
    xlsx_dir = root / "xlsx"
    out_dir = root / "json"
    xlsx_dir.mkdir()
    out_dir.mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"
    ws.append(["Key", "Name", "/stats/hp", "/stats/mp", "Tags"])
    ws.append(["WPN001", "Steel Sword", 10, 0, "melee;steel"])
    ws.append(["WPN002", "Silver Bow", 8, 5, "ranged;silver"])
    wb.save(xlsx_dir / "Items.xlsx")

    (root / "config.json").write_text(
        json.dumps(
            {
                "name": "Round",
                "environment": {"xlsx": "./xlsx", "out": "./json"},
                "xlsxtables": {
                    "Items.xlsx": {
                        "workSheets": [
                            {"name": "Equipment", "out": "Items_Equipment.json"},
                        ]
                    }
                },
            }
        ),
        encoding="utf-8",
    )


class RoundTripTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        _build_project(self.root)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_json_to_xlsx_preserves_data_through_full_loop(self) -> None:
        project = Project(self.root)

        # First export.
        project.export()
        json_path = self.root / "json" / "Items_Equipment.json"
        first = json.loads(json_path.read_text(encoding="utf-8"))

        # Round trip: json -> xlsx, then re-export xlsx -> json.
        result = project.import_json("items")
        self.assertEqual(result.errors, [])
        self.assertEqual(result.sheets, ["Equipment"])

        project.export()
        second = json.loads(json_path.read_text(encoding="utf-8"))

        self.assertEqual(first, second)

    def test_write_xlsx_preserves_other_sheets(self) -> None:
        xlsx_path = self.root / "xlsx" / "Items.xlsx"
        wb = load_workbook(xlsx_path)
        wb.create_sheet("Untouched").append(["leave_me", "alone"])
        wb.save(xlsx_path)

        ws = JSONWorksheet(
            source=xlsx_path,
            sheet_name="Equipment",
            columns=["Key", "Name"],
            rows=[{"Key": "X", "Name": "Y"}],
        )
        ws.write_xlsx(xlsx_path)

        reloaded = load_workbook(xlsx_path)
        self.assertIn("Untouched", reloaded.sheetnames)
        self.assertIn("Equipment", reloaded.sheetnames)
        self.assertEqual(reloaded["Untouched"]["A1"].value, "leave_me")

    def test_from_json_derives_columns_with_nesting(self) -> None:
        json_path = self.root / "rows.json"
        json_path.write_text(
            json.dumps(
                [
                    {"Key": "X", "stats": {"hp": 1, "mp": 2}, "Tags": ["a", "b"]},
                    {"Key": "Y", "stats": {"hp": 3, "mp": 4}, "Tags": ["c"]},
                ]
            ),
            encoding="utf-8",
        )
        ws = JSONWorksheet.from_json(json_path, sheet_name="S")
        self.assertEqual(ws.columns, ["Key", "/stats/hp", "/stats/mp", "Tags"])
        self.assertEqual(ws.sheet_name, "S")


if __name__ == "__main__":
    unittest.main()
