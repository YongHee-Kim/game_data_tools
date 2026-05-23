"""Context-managed wrapper around an openpyxl Workbook for batched JSON I/O.

Lets callers open an xlsx once and then read or write many sheets, instead of
re-opening (and re-saving) the file per sheet. ``Project.export`` /
``Project.import_json`` use this to avoid N opens for an N-sheet workbook;
``JSONWorksheet.read`` / ``write_xlsx`` are thin convenience wrappers for the
single-sheet case.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .worksheet import JSONWorksheet


class JSONWorkbook:
    """An open openpyxl workbook plus per-sheet JSON read/write helpers.

    Open with ``open_for_read`` (lazy/read-only) or ``open_for_write`` (full
    mode; creates a fresh workbook if the file does not exist). Use as a
    context manager so the underlying file handle is always released. For
    writes, call ``save()`` before exiting — the context manager only closes.
    """

    def __init__(self, wb: Workbook, source: Path):
        self._wb = wb
        self.source = source

    @classmethod
    def open_for_read(cls, xlsx_path: Path) -> "JSONWorkbook":
        return cls(load_workbook(xlsx_path, data_only=True, read_only=True), xlsx_path)

    @classmethod
    def open_for_write(cls, xlsx_path: Path) -> "JSONWorkbook":
        if xlsx_path.is_file():
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        return cls(wb, xlsx_path)

    def __enter__(self) -> "JSONWorkbook":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    @property
    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def read_sheet(self, sheet_name: str, **kwargs: Any) -> JSONWorksheet:
        """Read one sheet into a JSONWorksheet using this open workbook."""
        return JSONWorksheet._read_from_workbook(self._wb, self.source, sheet_name, **kwargs)

    def write_sheet(self, ws: JSONWorksheet, *, start_line: int = 1) -> None:
        """Replace (or add) a sheet inside this open workbook. Does not save."""
        ws._write_to_workbook(self._wb, start_line=start_line)

    def save(self) -> None:
        self._wb.save(self.source)

    def close(self) -> None:
        self._wb.close()
